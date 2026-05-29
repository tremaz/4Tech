"""
TrainSight API - Backend Python + FastAPI
Processa planilhas de treinamento e retorna métricas para Power Apps

Arquitetura:
  - Backend: Python + FastAPI (Azure App Service)
  - Frontend: Power Apps (Conector Personalizado)
  - Comunicação: HTTP/REST + JSON
"""

import json
import zipfile
import re
import io
from pathlib import Path
from typing import Optional
import xml.etree.ElementTree as ET
import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel

from models import (
    RespostaGupy,
    RespostaSupplyGo,
    RespostaCombinada,
    UploadUrlRequest,
    ResumoGupy,
    ResumoSupplyGo,
    Agrupamento,
)
from metrics import (
    detectar_fonte,
    processar_excel,
    calcular_metrics_gupy,
    calcular_metrics_supplygo,
    GROUPING_PATTERNS_GUPY,
    GROUPING_PATTERNS_SUPPLYGO,
)

# =============================================================================
# App FastAPI
# =============================================================================

app = FastAPI(
    title="TrainSight API",
    description="Processa planilhas de treinamento (Gupy/SupplyGo) e retorna métricas para Power Apps",
    version="1.0.0",
)

# CORS — permite que o Power Apps acesse a API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =============================================================================
# Carregar dados locais (para endpoints de consulta direta)
# =============================================================================

_data_gupy: Optional[pd.DataFrame] = None
_data_supplygo: Optional[pd.DataFrame] = None


def _load_gupy() -> pd.DataFrame:
    global _data_gupy
    if _data_gupy is None:
        try:
            _data_gupy = pd.read_excel(
                "./Participants-36809.xlsx",
                index_col=None,
                keep_default_na=False,
                na_values="nan",
            )
        except FileNotFoundError:
            pass
    return _data_gupy


def _load_supplygo() -> pd.DataFrame:
    global _data_supplygo
    if _data_supplygo is None:
        try:
            _data_supplygo = pd.read_excel(
                "./Relatório SupplyGo - Desafio_Hack-ta-on.xlsx",
                header=6,
                index_col=None,
                keep_default_na=False,
                na_values="nan",
                sheet_name="Relatorio",
            )
        except FileNotFoundError:
            pass
    return _data_supplygo


# =============================================================================
# Endpoints principais (conforme doc TrainSight)
# =============================================================================


@app.get("/")
def raiz():
    """Health check — retorna status da API."""
    gupy = _load_gupy()
    supplygo = _load_supplygo()
    return {
        "status": "online",
        "versao": "1.0.0",
        "records_gupy": len(gupy) if gupy is not None else 0,
        "records_supplygo": len(supplygo) if supplygo is not None else 0,
    }


@app.post("/upload")
async def upload_planilha(arquivo: UploadFile = File(...)):
    """
    Recebe o Excel, processa e retorna o dashboard em JSON.
    Detecta automaticamente se é Gupy ou SupplyGo.
    """
    try:
        conteudo = await arquivo.read()
        
        # Tenta ler o Excel com diferentes configurações
        df = None
        
        # Tenta ler com sheet_name=None para ver todas as abas
        try:
            xl = pd.ExcelFile(io.BytesIO(conteudo))
            # SupplyGo tem abas específicas
            if 'Relatorio' in xl.sheet_names:
                df = pd.read_excel(io.BytesIO(conteudo), header=6, sheet_name='Relatorio')
            elif 'Relatorio' in xl.sheet_names:
                df = pd.read_excel(io.BytesIO(conteudo), header=6, sheet_name='Relatorio')
            else:
                # Fallback: tenta ler primeira aba com header default
                df = pd.read_excel(io.BytesIO(conteudo))
        except Exception:
            # Fallback final
            df = pd.read_excel(io.BytesIO(conteudo))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler arquivo: {str(e)}")

    resultado = processar_excel(df)
    return resultado


@app.post("/upload-url")
async def upload_por_url(req: UploadUrlRequest):
    """
    Recebe URL do Excel, baixa e processa.
    Alternativa ao upload direto.
    """
    try:
        import requests
        response = requests.get(req.url, timeout=30)
        response.raise_for_status()
        df = pd.read_excel(io.BytesIO(response.content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao baixar/processar URL: {str(e)}")

    resultado = processar_excel(df)
    return resultado


# =============================================================================
# Endpoints de consulta direta (arquivos locais)
# =============================================================================


@app.get("/gupy/metrics", response_model=RespostaGupy)
def gupy_metrics():
    """Retorna métricas completas do Gupy (arquivo local)."""
    df = _load_gupy()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo Gupy não encontrado")
    resultado = calcular_metrics_gupy(df)
    return resultado


@app.get("/supplygo/metrics", response_model=RespostaSupplyGo)
def supplygo_metrics():
    """Retorna métricas completas do SupplyGo (arquivo local)."""
    df = _load_supplygo()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo SupplyGo não encontrado")
    resultado = calcular_metrics_supplygo(df)
    return resultado


@app.get("/combined/metrics")
def combined_metrics():
    """Retorna métricas combinadas de Gupy e SupplyGo."""
    gupy = _load_gupy()
    supplygo = _load_supplygo()

    gupy_data = {}
    if gupy is not None:
        gupy_data = calcular_metrics_gupy(gupy)

    supplygo_data = {}
    if supplygo is not None:
        supplygo_data = calcular_metrics_supplygo(supplygo)

    return {"gupy": gupy_data, "supplygo": supplygo_data}


# =============================================================================
# Endpoints de distribuição de status
# =============================================================================


@app.get("/gupy/status-distribution")
def gupy_status_distribution():
    """Retorna contagem de colaboradores por status (Gupy)."""
    df = _load_gupy()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo Gupy não encontrado")
    resultado = calcular_metrics_gupy(df)
    return resultado.get("status_distribution", {})


@app.get("/supplygo/status-distribution")
def supplygo_status_distribution():
    """Retorna contagem de colaboradores por status (SupplyGo)."""
    df = _load_supplygo()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo SupplyGo não encontrado")
    resultado = calcular_metrics_supplygo(df)
    return resultado.get("status_distribution", {})


# =============================================================================
# Endpoints de colaboradores (lista paginada)
# =============================================================================


@app.get("/gupy/colaboradores")
def gupy_colaboradores(
    estado: str = Query(None, description="Filtrar por estado (UF)"),
    status: str = Query(None, description="Filtrar por status"),
    pagina: int = Query(1, ge=1),
    limite: int = Query(50, ge=1, le=200),
):
    """Lista paginada de colaboradores do Gupy."""
    df = _load_gupy()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo Gupy não encontrado")

    if estado:
        df = df[df["Estado"] == estado]
    if status:
        df = df[df["Status de realização"] == status]

    total = len(df)
    inicio = (pagina - 1) * limite
    fim = inicio + limite
    paginado = df.iloc[inicio:fim]

    colunas = [
        "Nome", "Email", "Estado", "Cargo", "Status de realização",
        "Progresso", "Média", "Pontuação no ranking", "Data de início na trilha",
        "Data de finalização na trilha", "Área", "Unidade", "Diretoria",
    ]
    cols_exist = [c for c in colunas if c in df.columns]
    data = paginado[cols_exist].to_dict(orient="records")

    return {
        "total": total,
        "page": pagina,
        "page_size": limite,
        "total_pages": (total + limite - 1) // limite if limite else 1,
        "items": data,
    }


@app.get("/supplygo/colaboradores")
def supplygo_colaboradores(
    estado: str = Query(None, description="Filtrar por estado"),
    status: str = Query(None, description="Filtrar por status"),
    gerencia: str = Query(None, description="Filtrar por gerência"),
    pagina: int = Query(1, ge=1),
    limite: int = Query(50, ge=1, le=200),
):
    """Lista paginada de colaboradores do SupplyGo."""
    df = _load_supplygo()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo SupplyGo não encontrado")

    if estado:
        df = df[df["ESTADO"] == estado.upper()]
    if status:
        df = df[df["Status Trilha"] == status]
    if gerencia:
        df = df[df["GERÊNCIA CORPORATIVA"] == gerencia]

    total = len(df)
    inicio = (pagina - 1) * limite
    fim = inicio + limite
    paginado = df.iloc[inicio:fim]

    colunas = [
        "Nome", "CARGO", "FAMÍLIA DE CARGO", "ESTADO", "UNIDADE",
        "GERÊNCIA CORPORATIVA", "LIDERANÇA CORPORATIVA", "Status Trilha",
        "Carga Horária Total", "Carga Horária Cursada", "Problema na plataforma?",
    ]
    cols_exist = [c for c in colunas if c in df.columns]
    data = paginado[cols_exist].to_dict(orient="records")

    return {
        "total": total,
        "page": pagina,
        "page_size": limite,
        "total_pages": (total + limite - 1) // limite if limite else 1,
        "items": data,
    }


# =============================================================================
# Endpoints por gerência (SupplyGo)
# =============================================================================


@app.get("/supplygo/gerencias")
def supplygo_gerencias():
    """Retorna métricas por gerência corporativa (SupplyGo)."""
    df = _load_supplygo()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo SupplyGo não encontrado")

    gerencias = df.groupby("GERÊNCIA CORPORATIVA")
    resultado = []
    for nome, grupo in gerencias:
        total = len(grupo)
        conc = int((grupo["Status Trilha"] == "Concluído").sum())
        cursando = int((grupo["Status Trilha"] == "Cursando").sum())
        nao_inic = int((grupo["Status Trilha"] == "Não iniciado").sum())
        taxa = round(conc / total * 100, 1) if total else 0
        resultado.append({
            "nome": str(nome),
            "total": total,
            "concluido": conc,
            "cursando": cursando,
            "nao_iniciado": nao_inic,
            "taxa_conclusao": taxa,
        })
    return sorted(resultado, key=lambda x: x["total"], reverse=True)


@app.get("/supplygo/problemas")
def supplygo_problemas():
    """Retorna colaboradores que reportaram problemas na plataforma SupplyGo."""
    df = _load_supplygo()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo SupplyGo não encontrado")

    if "Problema na plataforma?" in df.columns:
        problemas = df[df["Problema na plataforma?"].fillna("").astype(str).str.strip() == "Sim"]
    elif "Problema na plataforma? " in df.columns:
        problemas = df[df["Problema na plataforma? "].fillna("").astype(str).str.strip() == "Sim"]
    else:
        problemas = pd.DataFrame()
    colunas = ["Nome", "ESTADO", "GERÊNCIA CORPORATIVA", "Status Trilha", "OBSERVAÇÃO"]
    cols_exist = [c for c in colunas if c in problemas.columns]
    data = problemas[cols_exist].to_dict(orient="records")

    return data


# =============================================================================
# Endpoints escaláveis - agrupamentos dinâmicos
# =============================================================================


@app.get("/gupy/grupos")
def gupy_grupos():
    """Retorna lista de todos os agrupamentos disponíveis no Gupy.
    
    Escalável: retorna qualquer coluna que esteja nos GROUPING_PATTERNS_GUPY
    """
    df = _load_gupy()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo Gupy não encontrado")
    
    # Processar para obter todos os agrupamentos
    resultado = calcular_metrics_gupy(df)
    
    # Retornar apenas os nomes dos agrupamentos disponíveis
    grupos = {}
    for key in resultado:
        if key.startswith("por_") and isinstance(resultado[key], list):
            grupos[key] = {
                "coluna": key.replace("por_", ""),
                "total_grupos": len(resultado[key]),
                "amostra": resultado[key][:3] if resultado[key] else []
            }
    
    return grupos


@app.get("/gupy/grupo/{nome_grupo}")
def gupy_grupo(nome_grupo: str):
    """Retorna dados de um agrupamento específico do Gupy.
    
    Escalável: suporta qualquer coluna dos GROUPING_PATTERNS_GUPY
    Ex: /gupy/grupo/por_estado, /gupy/grupo/por_subarea, /gupy/grupo/por_centro_custo
    """
    df = _load_gupy()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo Gupy não encontrado")
    
    resultado = calcular_metrics_gupy(df)
    
    if nome_grupo not in resultado:
        disponiveis = [k for k in resultado.keys() if k.startswith("por_")]
        raise HTTPException(
            status_code=404,
            detail=f"Grupo '{nome_grupo}' não encontrado. Disponíveis: {disponiveis}"
        )
    
    return resultado[nome_grupo]


@app.get("/supplygo/grupos")
def supplygo_grupos():
    """Retorna lista de todos os agrupamentos disponíveis no SupplyGo."""
    df = _load_supplygo()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo SupplyGo não encontrado")
    
    resultado = calcular_metrics_supplygo(df)
    
    grupos = {}
    for key in resultado:
        if key.startswith("por_") and isinstance(resultado[key], list):
            grupos[key] = {
                "coluna": key.replace("por_", ""),
                "total_grupos": len(resultado[key]),
                "amostra": resultado[key][:3] if resultado[key] else []
            }
    
    return grupos


@app.get("/supplygo/grupo/{nome_grupo}")
def supplygo_grupo(nome_grupo: str):
    """Retorna dados de um agrupamento específico do SupplyGo."""
    df = _load_supplygo()
    if df is None:
        raise HTTPException(status_code=500, detail="Arquivo SupplyGo não encontrado")
    
    resultado = calcular_metrics_supplygo(df)
    
    if nome_grupo not in resultado:
        disponiveis = [k for k in resultado.keys() if k.startswith("por_")]
        raise HTTPException(
            status_code=404,
            detail=f"Grupo '{nome_grupo}' não encontrado. Disponíveis: {disponiveis}"
        )
    
    return resultado[nome_grupo]


# =============================================================================
# Analyzer (mantido do main.py original — para engenharia reversa do Excel)
# =============================================================================


@app.get("/analyzer/structure")
def analyzer_structure(filename: str = Query("Relatorio-SupplyGo-Desafio_Hack-ta-on.xlsx")):
    """
    Analisa a estrutura interna de um arquivo Excel (pivot tables, KPIs, dependências).
    Mantido do script original de engenharia reversa.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado")

    xlsx_path = Path(filename)
    if not xlsx_path.exists():
        raise HTTPException(status_code=404, detail=f"Arquivo não encontrado: {filename}")

    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    wb_values = load_workbook(xlsx_path, data_only=True, keep_links=True)

    cell_ref_re = re.compile(
        r"(?:(?:'[^']+'|[A-Za-z0-9_]+)!)?\$?[A-Z]{1,3}\$?[0-9]{1,7}(?::\$?[A-Z]{1,3}\$?[0-9]{1,7})?"
    )

    def safe(v):
        return None if v is None else str(v)

    def get_non_empty_cells(ws, limit=200):
        arr = []
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    arr.append({
                        "celula": c.coordinate,
                        "valor": safe(c.value),
                        "tipo": c.data_type,
                    })
                    if len(arr) >= limit:
                        return arr
        return arr

    def detect_kpis(ws, limit_scan=800, max_items=50):
        non_empty = []
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    non_empty.append((c.coordinate, c.value))
                    if len(non_empty) >= limit_scan:
                        break
            if len(non_empty) >= limit_scan:
                break
        kpis = []
        for i, (coord, value) in enumerate(non_empty):
            sval = str(value).strip()
            numeric_like = re.fullmatch(
                r'[-+]?\d+[\.,]?\d*%?', sval
            ) or isinstance(value, (int, float))
            if numeric_like:
                prev_items = non_empty[max(0, i - 4):i]
                labels = [
                    str(v).strip()
                    for _, v in prev_items
                    if isinstance(v, str) and str(v).strip()
                ]
                if labels:
                    kpis.append({
                        "celula": coord,
                        "valor": sval,
                        "labels_proximos": labels[-4:],
                    })
            if len(kpis) >= max_items:
                break
        return kpis

    result = {
        "arquivo": filename,
        "abas": [],
        "kpis_por_aba": {},
        "pivot_tables": [],
        "botoes": [],
    }

    for ws in wb.worksheets:
        kpis = detect_kpis(wb_values[ws.title])
        result["kpis_por_aba"][ws.title] = kpis[:20]
        result["abas"].append({
            "nome": ws.title,
            "linhas": ws.max_row,
            "colunas": ws.max_column,
            "kpis_detectados": len(kpis),
        })

    with zipfile.ZipFile(xlsx_path, "r") as z:
        names = z.namelist()
        pivot_files = [n for n in names if n.startswith("xl/pivotTables/") and n.endswith(".xml")]
        for pf in pivot_files:
            root = ET.fromstring(z.read(pf))
            cache_id = root.attrib.get("cacheId")
            location = root.find(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}location")
            ref = location.attrib.get("ref") if location is not None else None
            result["pivot_tables"].append({
                "nome": root.attrib.get("name"),
                "cacheId": cache_id,
                "localizacao": ref,
                "arquivo_xml": pf,
            })

        buttons = []
        drawing_files = [n for n in names if n.startswith("xl/drawings/drawing") and n.endswith(".xml")]
        for df in drawing_files:
            xml_str = z.read(df).decode("utf-8", errors="ignore")
            texts = [t.strip() for t in re.findall(r"<a:t>(.*?)</a:t>", xml_str) if t.strip()]
            hlinks = re.findall(r'hlinkClick[^>]*r:id="(rId\d+)"', xml_str)
            if texts:
                buttons.append({"drawing": df, "textos": texts, "hyperlinks": len(hlinks)})
        result["botoes"] = buttons

    return result
